from __future__ import annotations

import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import PROJECT_ROOT

LOGGER = logging.getLogger(__name__)

EXCEL_FILE = Path(os.environ.get("EXCEL_FILE", PROJECT_ROOT / "bot" / "tickets_data.xlsx"))
GOOGLE_SHEETS_CREDENTIALS_FILE = Path(os.environ.get(
    "GOOGLE_SHEETS_CREDENTIALS_FILE",
    str(PROJECT_ROOT / "bot" / "google-service-account.json"),
))
if not GOOGLE_SHEETS_CREDENTIALS_FILE.is_absolute():
    GOOGLE_SHEETS_CREDENTIALS_FILE = PROJECT_ROOT / "bot" / GOOGLE_SHEETS_CREDENTIALS_FILE
GOOGLE_SHEETS_SPREADSHEET_ID = os.environ.get("GOOGLE_SHEETS_SPREADSHEET_ID", "")

HEADERS = [
    "ID билета",
    "Имя",
    "Фамилия",
    "Telegram ID",
    "Username",
    "Название мероприятия",
    "Дата покупки",
    "Статус",
    "Проход",
]
STATUS_CONFIRMED = "подтверждено"
CHECKED_IN_VALUE = "ПРОШЕЛ"
HEADER_COLOR = "061A0A"
HEADER_TEXT_COLOR = "7CFF8A"
DEFAULT_CELL_COLOR = "FFFFFF"
CHECKED_IN_COLOR = "C6EFCE"
GOOGLE_COLUMN_WIDTHS = [130, 150, 150, 130, 150, 220, 170, 170, 140]


def _format_date(value: str | None) -> str:
    if not value:
        return datetime.now().strftime("%d.%m.%Y %H:%M")
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).strftime("%d.%m.%Y %H:%M")
    except ValueError:
        return value


def _get_or_create_workbook():
    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    if EXCEL_FILE.exists():
        return load_workbook(EXCEL_FILE)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _style_header_row(sheet) -> None:
    for col_num, header in enumerate(HEADERS, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color=HEADER_TEXT_COLOR, name="Consolas")
        cell.fill = PatternFill("solid", start_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[get_column_letter(col_num)].width = 22


def _apply_special_row_styles(sheet) -> None:
    default_fill = PatternFill("solid", start_color=DEFAULT_CELL_COLOR, end_color=DEFAULT_CELL_COLOR)
    checked_in_fill = PatternFill("solid", start_color=CHECKED_IN_COLOR, end_color=CHECKED_IN_COLOR)
    for row_number in range(2, sheet.max_row + 1):
        for column_number in range(1, len(HEADERS) + 1):
            sheet.cell(row=row_number, column=column_number).fill = default_fill
        check_in_status = str(sheet.cell(row=row_number, column=9).value or "").strip().upper()
        if check_in_status == CHECKED_IN_VALUE:
            sheet.cell(row=row_number, column=9).fill = checked_in_fill


def _get_or_create_sheet(wb, event_name: str):
    sheet_name = (event_name or "Event")[:31]
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    sheet = wb.create_sheet(title=sheet_name)
    _style_header_row(sheet)
    return sheet


def _ticket_exists(wb, ticket_code: str) -> bool:
    normalized = str(ticket_code or "").strip().lstrip("#").upper()
    if not normalized:
        return False
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            existing = str(row[0] or "").strip().lstrip("#").upper()
            if existing == normalized:
                return True
    return False


def _sync_google_sheets() -> None:
    if not GOOGLE_SHEETS_SPREADSHEET_ID or not GOOGLE_SHEETS_CREDENTIALS_FILE.exists() or not EXCEL_FILE.exists():
        return

    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        LOGGER.warning("Google Sheets sync skipped: install gspread and google-auth")
        return

    def format_worksheet(worksheet, rows: list[list[str]]) -> None:
        row_count = max(len(rows), 1)
        col_count = max(len(rows[0]) if rows else len(HEADERS), 1)
        requests = [
            {
                "repeatCell": {
                    "range": {
                        "sheetId": worksheet.id,
                        "startRowIndex": 0,
                        "endRowIndex": 1,
                        "startColumnIndex": 0,
                        "endColumnIndex": col_count,
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "backgroundColor": {"red": 0.02, "green": 0.10, "blue": 0.04},
                            "horizontalAlignment": "CENTER",
                            "textFormat": {
                                "bold": True,
                                "foregroundColor": {"red": 0.49, "green": 1, "blue": 0.54},
                                "fontSize": 11,
                                "fontFamily": "Consolas",
                            },
                        }
                    },
                    "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)",
                }
            },
            {
                "repeatCell": {
                    "range": {
                        "sheetId": worksheet.id,
                        "startRowIndex": 1,
                        "endRowIndex": max(row_count, 2),
                        "startColumnIndex": 0,
                        "endColumnIndex": col_count,
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "backgroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
                            "verticalAlignment": "MIDDLE",
                            "wrapStrategy": "WRAP",
                        }
                    },
                    "fields": "userEnteredFormat(backgroundColor,verticalAlignment,wrapStrategy)",
                }
            },
            {
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": worksheet.id,
                        "gridProperties": {"frozenRowCount": 1},
                    },
                    "fields": "gridProperties.frozenRowCount",
                }
            },
            {
                "setBasicFilter": {
                    "filter": {
                        "range": {
                            "sheetId": worksheet.id,
                            "startRowIndex": 0,
                            "endRowIndex": max(row_count, 1),
                            "startColumnIndex": 0,
                            "endColumnIndex": col_count,
                        }
                    }
                }
            },
        ]

        for column_index, width in enumerate(GOOGLE_COLUMN_WIDTHS[: min(col_count, len(GOOGLE_COLUMN_WIDTHS))]):
            requests.append(
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": worksheet.id,
                            "dimension": "COLUMNS",
                            "startIndex": column_index,
                            "endIndex": column_index + 1,
                        },
                        "properties": {"pixelSize": width},
                        "fields": "pixelSize",
                    }
                }
            )

        for row_index, row in enumerate(rows[1:], start=1):
            check_in_value = (row[8] if len(row) > 8 else "").strip().upper()
            if check_in_value != CHECKED_IN_VALUE:
                continue
            requests.append(
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": worksheet.id,
                            "startRowIndex": row_index,
                            "endRowIndex": row_index + 1,
                            "startColumnIndex": 8,
                            "endColumnIndex": 9,
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.776, "green": 0.937, "blue": 0.808}
                            }
                        },
                        "fields": "userEnteredFormat.backgroundColor",
                    }
                }
            )

        spreadsheet.batch_update({"requests": requests})

    try:
        credentials = Credentials.from_service_account_file(
            str(GOOGLE_SHEETS_CREDENTIALS_FILE),
            scopes=[
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive",
            ],
        )
        client = gspread.authorize(credentials)
        spreadsheet = client.open_by_key(GOOGLE_SHEETS_SPREADSHEET_ID)
        workbook = load_workbook(EXCEL_FILE)
        workbook_sheet_titles = {sheet.title[:100] for sheet in workbook.worksheets}

        for sheet in workbook.worksheets:
            rows = [
                ["" if cell is None else str(cell) for cell in row]
                for row in sheet.iter_rows(values_only=True)
            ]
            if not rows:
                continue

            title = sheet.title[:100]
            try:
                worksheet = spreadsheet.worksheet(title)
                worksheet.clear()
            except gspread.WorksheetNotFound:
                worksheet = spreadsheet.add_worksheet(title=title, rows=max(len(rows), 1), cols=max(len(rows[0]), 1))

            worksheet.resize(rows=max(len(rows), 2), cols=max(len(rows[0]), 1))
            worksheet.update("A1", rows, value_input_option="USER_ENTERED")
            format_worksheet(worksheet, rows)

        for worksheet in spreadsheet.worksheets():
            if worksheet.title not in workbook_sheet_titles and len(spreadsheet.worksheets()) > 1:
                spreadsheet.del_worksheet(worksheet)
    except Exception as exc:
        LOGGER.warning("Google Sheets sync failed: %s", exc)


def _save_workbook(wb) -> None:
    wb.save(EXCEL_FILE)
    _sync_google_sheets()


def save_confirmed_order(conn: Any, order: dict[str, Any]) -> None:
    event = conn.events.find_one({"id": order["event_id"]})
    if not event:
        return

    tickets = list(conn.tickets.find({"order_id": order["id"]}).sort("id", 1))
    if not tickets:
        return

    wb = _get_or_create_workbook()
    sheet = _get_or_create_sheet(wb, event["title"])
    _style_header_row(sheet)
    saved_any = False

    for ticket in tickets:
        ticket_code = ticket.get("ticket_code")
        if _ticket_exists(wb, ticket_code):
            continue
        sheet.append(
            [
                ticket_code,
                order.get("customer_first_name") or "",
                order.get("customer_last_name") or "",
                order.get("telegram_id") or "",
                order.get("telegram_username") or order.get("customer_phone") or order.get("customer_email") or "",
                event.get("title") or "",
                _format_date(order.get("created_at")),
                STATUS_CONFIRMED,
                "",
            ]
        )
        saved_any = True

    if not saved_any:
        return
    _apply_special_row_styles(sheet)
    _save_workbook(wb)


def mark_ticket_checked_in(ticket_code: str) -> None:
    normalized = str(ticket_code or "").strip().lstrip("#").upper()
    if not normalized or not EXCEL_FILE.exists():
        return

    wb = load_workbook(EXCEL_FILE)
    changed = False
    for sheet in wb.worksheets:
        for row_number in range(2, sheet.max_row + 1):
            existing = str(sheet.cell(row=row_number, column=1).value or "").strip().lstrip("#").upper()
            if existing != normalized:
                continue
            _style_header_row(sheet)
            sheet.cell(row=row_number, column=9).value = CHECKED_IN_VALUE
            _apply_special_row_styles(sheet)
            changed = True
            break

    if changed:
        _save_workbook(wb)
